import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Fix encoding for Windows console
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# Path to the FW_ data Base folder
base_path = "FW_ data Base"

# List of files
files = [
    "14th oct - docusign.xlsx",
    "15th oct- intel.xlsx",
    "16th Oct.xlsx"
]

print("=" * 80)
print("COLUMN NAMES COMPARISON FROM 3 EXCEL FILES")
print("=" * 80)
print()

# Dictionary to store column names for each file
file_columns = {}

# Read column names from each file
for file_name in files:
    file_path = os.path.join(base_path, file_name)
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        continue
    
    try:
        print(f"📄 Reading: {file_name}")
        
        # Try different engines for reading Excel files
        try:
            # Try openpyxl first (for .xlsx)
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
        except Exception:
            try:
                # Try xlrd for old .xls files
                excel_file = pd.ExcelFile(file_path, engine='xlrd')
            except Exception:
                # Try default engine
                excel_file = pd.ExcelFile(file_path)
        
        sheet_name = excel_file.sheet_names[0]  # Get first sheet
        
        # First, read with pandas normally
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', header=0, nrows=0)
            engine_used = 'openpyxl'
        except Exception:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd', header=0, nrows=0)
                engine_used = 'xlrd'
            except Exception:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=0, nrows=0)
                engine_used = 'default'
        
        # Get columns from pandas
        columns_pd = []
        for col in df.columns:
            col_str = str(col).strip()
            if col_str == '' or pd.isna(col) or 'Unnamed' in col_str:
                columns_pd.append(f'Column_{len(columns_pd)+1}')
            else:
                columns_pd.append(col_str)
        
        # Try to get column count from openpyxl if possible
        max_col = len(columns_pd)
        columns = columns_pd.copy()
        
        try:
            # Also check actual column count from openpyxl
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name]
            max_col_excel = ws.max_column
            
            if max_col_excel > max_col:
                # Read header row directly from openpyxl to get all column names
                columns_excel = []
                for col_idx in range(1, max_col_excel + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    if cell_value is None or str(cell_value).strip() == '':
                        columns_excel.append(f'Column_{col_idx}')
                    else:
                        columns_excel.append(str(cell_value).strip())
                
                # Use Excel columns if we got more
                if len(columns_excel) > len(columns):
                    columns = columns_excel
                    max_col = max_col_excel
            
            wb.close()
        except Exception:
            # If openpyxl doesn't work, use pandas columns
            pass
        
        file_columns[file_name] = {
            'columns': columns,
            'count': len(columns),
            'max_col': max_col,
            'sheet': sheet_name
        }
        
        print(f"   ✅ Sheet: {sheet_name}")
        print(f"   ✅ Columns: {len(columns)} (Excel max_col: {max_col})")
        print()
        
        wb.close()
        
    except Exception as e:
        print(f"   ❌ Error reading {file_name}: {str(e)}")
        print()

print("=" * 80)
print("COLUMN NAMES DETAILS")
print("=" * 80)
print()

# Display all column names for each file
for file_name, data in file_columns.items():
    print(f"\n📄 {file_name}")
    print(f"   Sheet: {data['sheet']}")
    print(f"   Total Columns: {data['count']}")
    print(f"   Column Names:")
    for idx, col in enumerate(data['columns'], 1):
        print(f"      {idx:2d}. {col}")

print()
print("=" * 80)
print("COLUMN COMPARISON")
print("=" * 80)
print()

# Compare columns across files
if len(file_columns) >= 2:
    file_names = list(file_columns.keys())
    
    # Get column sets for comparison
    column_sets = {}
    for file_name, data in file_columns.items():
        column_sets[file_name] = set([col.strip() for col in data['columns']])
    
    # Check if all sets are identical
    first_file = file_names[0]
    first_set = column_sets[first_file]
    
    all_match = True
    for i in range(1, len(file_names)):
        current_file = file_names[i]
        current_set = column_sets[current_file]
        
        if first_set == current_set:
            print(f"✅ {first_file} and {current_file} have IDENTICAL columns")
        else:
            all_match = False
            print(f"❌ {first_file} and {current_file} have DIFFERENT columns")
            
            # Show differences
            only_in_first = first_set - current_set
            only_in_current = current_set - first_set
            common = first_set & current_set
            
            if only_in_first:
                print(f"   Columns only in {first_file}: {sorted(only_in_first)}")
            if only_in_current:
                print(f"   Columns only in {current_file}: {sorted(only_in_current)}")
            print(f"   Common columns: {len(common)}/{len(first_set | current_set)}")
        print()
    
    if all_match:
        print("✅ ALL FILES HAVE IDENTICAL COLUMN NAMES!")
        print(f"   Common column count: {len(first_set)}")
        print(f"   Column names: {sorted(first_set)}")
    else:
        print("⚠️ FILES HAVE DIFFERENT COLUMN NAMES!")
        
        # Show common columns
        if len(file_columns) >= 2:
            common_cols = set.intersection(*column_sets.values())
            if common_cols:
                print(f"\n   Common columns across all files ({len(common_cols)}):")
                for col in sorted(common_cols):
                    print(f"      - {col}")
            
            # Show all unique columns
            all_cols = set.union(*column_sets.values())
            print(f"\n   All unique columns across all files ({len(all_cols)}):")
            for col in sorted(all_cols):
                files_with_col = [fname for fname, cols in column_sets.items() if col in cols]
                print(f"      - {col} (in: {', '.join(files_with_col)})")

print("=" * 80)

